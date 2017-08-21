import re
import csv
import math
import pandas as pd
import json
from openpyxl import Workbook
import os
from datetime import datetime
from openpyxl import load_workbook

all_resource_path = os.path.dirname(__file__)
today_path = os.path.join(all_resource_path, datetime.now().strftime('%y-%m-%d'))


def readcsv():
    df = pd.read_csv('uevent.csv')
    df = df['页面']
    l = [i for i in df.values]


def read_json_and_insert_xlsx():
    with open('2amzdemo.json') as json_file:
        data = json.load(json_file)
        json_dict = data
        arr = json_dict['data']
        for item in arr:
            insert_item_to_xlsx(item)


def insert_item_to_xlsx(item):
    category = item['category']
    line = [category, item['sub_category'], item['title'], item['rank']]  # 把数据中每一项整理出来
    belong_dir = os.path.join(today_path, item['belong'])
    if not os.path.exists(belong_dir):
        pass
        # os.makedirs(belong_dir)

    wb, xlsx_path = create_wb(item)

    ws = wb.active

    if category in wb.sheetnames:
        ws = wb.get_sheet_by_name(category)
        ws.append(line)

    # 创建新的sheet
    if category not in wb.sheetnames:
        if ws.title == 'Sheet':
            ws.title = category
        else:
            ws = wb.create_sheet(title=category)
        ws.append(['分类', '子分类', '标题', '排名'])  # 设置表头
        ws.append(line)  # 将数据以行的形式添加到xlsx中

    wb.save(xlsx_path)  # 保存xlsx文件


def create_wb(item):
    xlsx_path = os.path.join(today_path, item['category'] + '.xlsx')
    if os.path.exists(xlsx_path):
        wb1 = load_workbook(filename=xlsx_path)
    else:
        wb1 = Workbook()

    return wb1, xlsx_path


if __name__ == '__main__':
    read_json_and_insert_xlsx()
